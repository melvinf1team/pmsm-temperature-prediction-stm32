"""Generate the two Excel inventories for the PMSM datalogging datasets.

The workbooks are intentionally self-documenting: measured values, inferred
metadata and fields that still need a human confirmation are kept separate.
Existing values entered in the manual columns are preserved when the script is
run again.
"""

from __future__ import annotations

import hashlib
import re
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RAW_DIR = ROOT / "datalogging" / "logs"
PROCESSED_DIR = ROOT / "pretraitement" / "logs_processed_ewma"
RAW_WORKBOOK = ROOT / "inventaire_donnees_brutes.xlsx"
PROCESSED_WORKBOOK = ROOT / "inventaire_donnees_pretraitees_ewma.xlsx"

RAW_COLUMNS = [
    "stm32_time_ms",
    "d6t_temp_c",
    "ds18b20_temp_c",
    "motor_ud_v",
    "motor_uq_v",
    "motor_speed_mech_rpm",
    "motor_id_a",
    "motor_iq_a",
]

FEATURE_INPUT_COLUMNS = [
    "ds18b20_temp_c",
    "motor_ud_v",
    "motor_uq_v",
    "motor_speed_mech_rpm",
    "motor_id_a",
    "motor_iq_a",
]

DERIVED_COLUMNS = ["u_s", "i_s", "S_el", "speed_current", "speed_power"]
EWM_COLUMNS = FEATURE_INPUT_COLUMNS + DERIVED_COLUMNS
REFERENCE_SPANS = [1320, 3360, 6360, 9480]
REFERENCE_FREQUENCY_HZ = 2.0
ACTIVE_SPEED_THRESHOLD_RPM = 1500.0
SESSION_RE = re.compile(
    r"^daq_log_(?:(?P<brake>\d+(?:\.\d+)?)_)?"
    r"(?P<date>\d{8})_(?P<time>\d{6})\.csv$"
)

MANUAL_COLUMNS = [
    "Opérateur",
    "Banc / moteur",
    "Objet de l'essai (équipe)",
    "Commentaires équipe",
    "Statut de validation",
    "Mise à jour manuelle",
]

# Professional palette that remains readable in Excel Online and when printed.
NAVY = "003B5C"
BLUE = "0077B6"
CYAN = "00A3E0"
GREEN = "49A942"
DARK_GREEN = "2E7D32"
AMBER = "F4B183"
YELLOW = "FFF2CC"
RED = "C00000"
LIGHT_BLUE = "DDEBF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9E1F2"
WHITE = "FFFFFF"
BLACK = "1F1F1F"
THIN_GRAY = Side(style="thin", color="D9E1F2")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def git_value(*args: str) -> str:
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip()
    except (FileNotFoundError, subprocess.CalledProcessError):
        return "Non disponible"


def git_file_info(path: Path) -> tuple[str, str]:
    relative = path.relative_to(ROOT).as_posix()
    value = git_value(
        "log",
        "-1",
        "--format=%h|%ad",
        "--date=iso-strict",
        "--",
        relative,
    )
    if "|" in value:
        return tuple(value.split("|", 1))  # type: ignore[return-value]
    return value, ""


def parse_session_name(path: Path) -> dict[str, Any]:
    match = SESSION_RE.match(path.name)
    if not match:
        raise ValueError(f"Nom de fichier non reconnu : {path.name}")
    stamp = f"{match.group('date')}_{match.group('time')}"
    nominal_start = datetime.strptime(stamp, "%Y%m%d_%H%M%S")
    brake_text = match.group("brake")
    return {
        "session_id": stamp,
        "nominal_start": nominal_start,
        "brake_value": float(brake_text) if brake_text else None,
    }


def duration_value(seconds: float) -> float:
    """Return an Excel duration value (fraction of a day)."""
    return seconds / 86400.0


def duration_text(seconds: float) -> str:
    milliseconds = int(round(seconds * 1000.0))
    hours, remainder = divmod(milliseconds, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    secs, millis = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}.{millis:03d}"


def add_physical_features(frame: pd.DataFrame) -> None:
    frame["u_s"] = np.hypot(frame["motor_ud_v"], frame["motor_uq_v"])
    frame["i_s"] = np.hypot(frame["motor_id_a"], frame["motor_iq_a"])
    frame["S_el"] = 1.5 * frame["u_s"] * frame["i_s"]
    frame["speed_current"] = frame["motor_speed_mech_rpm"] * frame["i_s"]
    frame["speed_power"] = frame["motor_speed_mech_rpm"] * frame["S_el"]


def scaled_spans(frequency_hz: float) -> list[int]:
    return [
        max(1, int(round(span * frequency_hz / REFERENCE_FREQUENCY_HZ)))
        for span in REFERENCE_SPANS
    ]


def expected_processed_frame(raw_frame: pd.DataFrame, spans: list[int]) -> pd.DataFrame:
    frame = raw_frame.copy()
    add_physical_features(frame)
    ewma: dict[str, pd.Series] = {}
    for column in EWM_COLUMNS:
        for span in spans:
            ewma[f"{column}_ewma_{span}"] = frame[column].ewm(
                span=span, adjust=False
            ).mean()
    frame = pd.concat([frame, pd.DataFrame(ewma, index=frame.index)], axis=1)
    output_columns = ["d6t_temp_c"]
    for column in EWM_COLUMNS:
        output_columns.append(column)
        output_columns.extend(f"{column}_ewma_{span}" for span in spans)
    result = frame[output_columns].replace([np.inf, -np.inf], np.nan).fillna(0.0)
    return result


def raw_run_segments(
    frame: pd.DataFrame, nominal_start: datetime
) -> list[dict[str, Any]]:
    """Return contiguous motor-on phases of at least ten seconds.

    Gaps shorter than five seconds are merged to ignore isolated estimator
    drop-outs. Longer interruptions remain separate sessions.
    """
    time_s = frame["stm32_time_ms"].to_numpy(dtype=float) / 1000.0
    offset_s = time_s - time_s[0]
    speed = frame["motor_speed_mech_rpm"].to_numpy(dtype=float)
    active = np.abs(speed) > ACTIVE_SPEED_THRESHOLD_RPM
    changes = np.flatnonzero(active[1:] != active[:-1]) + 1
    boundaries = np.concatenate(([0], changes, [len(active)]))
    ranges: list[list[int]] = []
    for start, stop in zip(boundaries[:-1], boundaries[1:]):
        if active[start]:
            ranges.append([int(start), int(stop - 1)])

    median_period_s = float(np.median(np.diff(time_s)))
    merged: list[list[int]] = []
    for start, stop in ranges:
        if merged:
            gap_s = offset_s[start] - offset_s[merged[-1][1]]
            if gap_s <= 5.0:
                merged[-1][1] = stop
                continue
        merged.append([start, stop])

    result: list[dict[str, Any]] = []
    for start, stop in merged:
        phase_active = active[start : stop + 1]
        active_count = int(phase_active.sum())
        active_duration = active_count * median_period_s
        if active_duration < 10.0:
            continue
        indices = np.arange(start, stop + 1)[phase_active]
        speeds = speed[indices]
        iq = frame["motor_iq_a"].to_numpy(dtype=float)[indices]
        start_offset = float(offset_s[start])
        end_offset = float(offset_s[stop])
        result.append(
            {
                "sequence": len(result) + 1,
                "start_offset_s": start_offset,
                "end_offset_s": end_offset,
                "duration_s": active_duration,
                "nominal_start": nominal_start + timedelta(seconds=start_offset),
                "nominal_end": nominal_start + timedelta(seconds=end_offset),
                "speed_mean": float(np.mean(speeds)),
                "speed_median": float(np.median(speeds)),
                "speed_p05": float(np.quantile(speeds, 0.05)),
                "speed_p95": float(np.quantile(speeds, 0.95)),
                "iq_mean": float(np.mean(iq)),
                "iq_median": float(np.median(iq)),
                "iq_p05": float(np.quantile(iq, 0.05)),
                "iq_p95": float(np.quantile(iq, 0.95)),
                "d6_start": float(frame["d6t_temp_c"].iloc[start]),
                "d6_end": float(frame["d6t_temp_c"].iloc[stop]),
                "d6_max": float(frame["d6t_temp_c"].iloc[start : stop + 1].max()),
                "ds_start": float(frame["ds18b20_temp_c"].iloc[start]),
                "ds_end": float(frame["ds18b20_temp_c"].iloc[stop]),
            }
        )
    return result


def quality_label(
    dt_ms: pd.Series, missing: int, infinite: int, duplicate_rows: int
) -> tuple[str, str]:
    nonpositive = int((dt_ms <= 0).sum())
    if missing or infinite or duplicate_rows or nonpositive:
        return (
            "À contrôler",
            f"manquantes={missing}, infinies={infinite}, doublons={duplicate_rows}, "
            f"intervalles non positifs={nonpositive}",
        )
    median = float(dt_ms.median())
    minimum = float(dt_ms.min())
    maximum = float(dt_ms.max())
    gaps = int((dt_ms > 2.0 * median).sum())
    if gaps:
        return "À contrôler", f"{gaps} intervalle(s) > 2 fois la période médiane"
    if minimum == maximum == median:
        return "OK", "Cadence uniforme"
    return "OK", f"Cadence moyenne intacte ; jitter {minimum:g}–{maximum:g} ms"


def discover_processed_files() -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for path in sorted(PROCESSED_DIR.glob("*.csv")):
        parsed = parse_session_name(path)
        session_id = parsed["session_id"]
        if session_id in result:
            raise ValueError(f"Plusieurs fichiers prétraités pour {session_id}")
        result[session_id] = {"path": path, **parsed}
    return result


def brake_metadata(
    session_id: str, processed_by_session: dict[str, dict[str, Any]]
) -> dict[str, Any]:
    processed = processed_by_session.get(session_id)
    if processed and processed["brake_value"] is not None:
        value = float(processed["brake_value"])
        return {
            "brake_mode": "Fixe",
            "brake_value": value,
            "brake_label": f"{value:.2f} A".replace(".", ","),
            "brake_source": "Préfixe du nom du fichier prétraité",
            "brake_confidence": "Élevée (nom), à valider",
        }
    return {
        "brake_mode": "Mixte",
        "brake_value": None,
        "brake_label": "Mixte — plusieurs niveaux",
        "brake_source": "Inférence : nom sans préfixe + paliers d'Iq moteur",
        "brake_confidence": "À confirmer par l'équipe",
    }


def analyse_raw_files() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    processed_by_session = discover_processed_files()
    records: list[dict[str, Any]] = []
    all_sequences: list[dict[str, Any]] = []

    for path in sorted(RAW_DIR.glob("*.csv")):
        parsed = parse_session_name(path)
        session_id = parsed["session_id"]
        frame = pd.read_csv(path, sep=";")
        if list(frame.columns) != RAW_COLUMNS:
            raise ValueError(f"Schéma brut inattendu dans {path.name}: {list(frame.columns)}")
        numeric = frame.apply(pd.to_numeric, errors="coerce")
        time_ms = numeric["stm32_time_ms"]
        dt_ms = time_ms.diff().dropna()
        positive_dt = dt_ms[dt_ms > 0]
        if positive_dt.empty:
            raise ValueError(f"Aucun intervalle temporel valide dans {path.name}")
        period_ms = float(positive_dt.median())
        frequency_hz = 1000.0 / period_ms
        duration_s = float((time_ms.iloc[-1] - time_ms.iloc[0]) / 1000.0)
        nominal_start = parsed["nominal_start"]
        nominal_end = nominal_start + timedelta(seconds=duration_s)

        array = numeric.to_numpy(dtype=float)
        missing = int(np.isnan(array).sum())
        infinite = int(np.isinf(array).sum())
        duplicate_rows = int(numeric.duplicated().sum())
        quality, quality_detail = quality_label(
            dt_ms, missing, infinite, duplicate_rows
        )

        active = numeric["motor_speed_mech_rpm"].abs() > ACTIVE_SPEED_THRESHOLD_RPM
        active_speed = numeric.loc[active, "motor_speed_mech_rpm"]
        active_iq = numeric.loc[active, "motor_iq_a"]
        if active_speed.empty:
            speed_setpoint = None
            speed_source = "Non déterminable"
            speed_profile = "Moteur à l'arrêt"
            speed_values = {key: None for key in ["median", "p05", "p95", "max"]}
            iq_values = {key: None for key in ["median", "p05", "p95"]}
        else:
            median_speed = float(active_speed.median())
            speed_setpoint = int(round(median_speed / 100.0) * 100)
            speed_source = "Inférée du plateau de vitesse mesurée"
            speed_profile = f"Arrêt ↔ {speed_setpoint:,} tr/min".replace(",", " ")
            speed_values = {
                "median": median_speed,
                "p05": float(active_speed.quantile(0.05)),
                "p95": float(active_speed.quantile(0.95)),
                "max": float(active_speed.max()),
            }
            iq_values = {
                "median": float(active_iq.median()),
                "p05": float(active_iq.quantile(0.05)),
                "p95": float(active_iq.quantile(0.95)),
            }

        active_duration_s = float(active.sum() * period_ms / 1000.0)
        active_ratio = float(active.mean() * 100.0)
        sequences = raw_run_segments(numeric, nominal_start)
        brake = brake_metadata(session_id, processed_by_session)

        if len(sequences) > 1:
            thermal_profile = f"Cycles de chauffe/refroidissement ({len(sequences)} marches)"
        elif sequences:
            trailing_s = duration_s - sequences[-1]["end_offset_s"]
            if trailing_s >= 120.0:
                thermal_profile = "Chauffe puis refroidissement"
            elif active.iloc[-1]:
                thermal_profile = "Chauffe ; refroidissement final non enregistré"
            else:
                thermal_profile = "Chauffe ; refroidissement final court"
        else:
            thermal_profile = "Acquisition moteur à l'arrêt"

        processed = processed_by_session.get(session_id)
        pair_name = processed["path"].name if processed else "Absent"
        pair_status = "OK — apparié par ID session" if processed else "Manquant"

        notes = [quality_detail]
        if brake["brake_mode"] == "Mixte":
            notes.append("Frein mixte inféré ; le courant de frein n'est pas télémesuré")
        if len(sequences) > 1:
            notes.append(f"{len(sequences)} phases moteur principales")
        if active.iloc[-1]:
            notes.append("Fichier terminé moteur en marche")
        if speed_values["max"] and speed_setpoint and speed_values["max"] > 1.10 * speed_setpoint:
            notes.append(f"Pointe transitoire à {speed_values['max']:.1f} tr/min")

        if brake["brake_mode"] == "Fixe":
            brake_description = f"frein {brake['brake_label']}"
        else:
            brake_description = "profil de frein mixte (à confirmer)"
        if active.iloc[-1]:
            ending_description = "acquisition arrêtée pendant la marche"
        elif len(sequences) > 1:
            ending_description = f"{len(sequences)} phases de marche avec refroidissements"
        else:
            ending_description = "phase de marche suivie d'un refroidissement"
        description = (
            f"Essai PMSM à {speed_setpoint:,} tr/min, {brake_description} ; "
            f"{ending_description}."
        ).replace(",", " ", 1) if speed_setpoint else "Acquisition sans marche moteur détectée."

        git_commit, git_date = git_file_info(path)
        record = {
            "session_id": session_id,
            "file_name": path.name,
            "path": path,
            "relative_path": path.relative_to(ROOT).as_posix(),
            "nominal_start": nominal_start,
            "nominal_end": nominal_end,
            "duration_s": duration_s,
            "rows": len(numeric),
            "columns": len(numeric.columns),
            "frequency_hz": frequency_hz,
            "period_ms": period_ms,
            "period_min_ms": float(dt_ms.min()),
            "period_max_ms": float(dt_ms.max()),
            "time_gaps": int((dt_ms > 2.0 * period_ms).sum()),
            "nonpositive_intervals": int((dt_ms <= 0).sum()),
            "missing": missing,
            "infinite": infinite,
            "duplicate_rows": duplicate_rows,
            "quality": quality,
            "quality_detail": quality_detail,
            "speed_setpoint": speed_setpoint,
            "speed_source": speed_source,
            "speed_profile": speed_profile,
            "speed_median": speed_values["median"],
            "speed_p05": speed_values["p05"],
            "speed_p95": speed_values["p95"],
            "speed_max": speed_values["max"],
            "active_duration_s": active_duration_s,
            "active_ratio": active_ratio,
            "sequence_count": len(sequences),
            "iq_median": iq_values["median"],
            "iq_p05": iq_values["p05"],
            "iq_p95": iq_values["p95"],
            "d6_start": float(numeric["d6t_temp_c"].iloc[0]),
            "d6_min": float(numeric["d6t_temp_c"].min()),
            "d6_max": float(numeric["d6t_temp_c"].max()),
            "d6_end": float(numeric["d6t_temp_c"].iloc[-1]),
            "d6_range": float(numeric["d6t_temp_c"].max() - numeric["d6t_temp_c"].min()),
            "ds_start": float(numeric["ds18b20_temp_c"].iloc[0]),
            "ds_min": float(numeric["ds18b20_temp_c"].min()),
            "ds_max": float(numeric["ds18b20_temp_c"].max()),
            "ds_end": float(numeric["ds18b20_temp_c"].iloc[-1]),
            "ds_range": float(
                numeric["ds18b20_temp_c"].max() - numeric["ds18b20_temp_c"].min()
            ),
            "thermal_profile": thermal_profile,
            "processed_name": pair_name,
            "pair_status": pair_status,
            "description": description,
            "notes": " | ".join(notes),
            "first_stm32_ms": float(time_ms.iloc[0]),
            "last_stm32_ms": float(time_ms.iloc[-1]),
            "size_bytes": path.stat().st_size,
            "sha256": sha256_file(path),
            "git_commit": git_commit,
            "git_date": git_date,
            **brake,
        }
        records.append(record)
        for sequence in sequences:
            all_sequences.append(
                {
                    "session_id": session_id,
                    "file_name": path.name,
                    **brake,
                    **sequence,
                }
            )

    records.sort(key=lambda item: item["nominal_start"])
    all_sequences.sort(key=lambda item: (item["session_id"], item["sequence"]))
    return records, all_sequences


def analyse_processed_files(
    raw_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    raw_by_session = {record["session_id"]: record for record in raw_records}
    records: list[dict[str, Any]] = []
    for path in sorted(PROCESSED_DIR.glob("*.csv")):
        parsed = parse_session_name(path)
        session_id = parsed["session_id"]
        raw = raw_by_session.get(session_id)
        if not raw:
            raise ValueError(f"Aucun brut source trouvé pour {path.name}")

        processed = pd.read_csv(path, sep=";", header=None)
        values = processed.to_numpy(dtype=float)
        missing = int(np.isnan(values).sum())
        infinite = int(np.isinf(values).sum())
        malformed = int(processed.shape[1] != 56)

        raw_frame = pd.read_csv(raw["path"], sep=";").apply(
            pd.to_numeric, errors="coerce"
        )
        spans = scaled_spans(raw["frequency_hz"])
        expected = expected_processed_frame(raw_frame, spans).to_numpy(dtype=float)
        shape_match = values.shape == expected.shape
        if shape_match:
            differences = np.abs(values - expected)
            max_abs_difference = float(differences.max(initial=0.0))
            divergent_cells = int((differences > 5.1e-7).sum())
        else:
            max_abs_difference = None
            divergent_cells = None

        rows_match = len(processed) == raw["rows"]
        finite_ok = missing == 0 and infinite == 0
        full_ok = (
            shape_match
            and rows_match
            and finite_ok
            and divergent_cells == 0
            and not malformed
        )
        validation = (
            "OK — conforme au prétraitement de référence (arrondi 10⁻⁶)"
            if full_ok
            else "À contrôler"
        )
        git_commit, git_date = git_file_info(path)

        records.append(
            {
                **raw,
                "file_name": path.name,
                "path": path,
                "relative_path": path.relative_to(ROOT).as_posix(),
                "raw_name": raw["file_name"],
                "raw_path": raw["path"],
                "raw_rows": raw["rows"],
                "raw_sha256": raw["sha256"],
                "rows": len(processed),
                "columns": processed.shape[1],
                "target_columns": 1,
                "feature_columns": processed.shape[1] - 1,
                "header": "Non",
                "time_included": "Non — hérité du brut uniquement",
                "separator": ";",
                "decimal": ".",
                "precision": "6 décimales",
                "spans": spans,
                "rows_match": rows_match,
                "shape_match": shape_match,
                "missing": missing,
                "infinite": infinite,
                "max_abs_difference": max_abs_difference,
                "divergent_cells": divergent_cells,
                "validation": validation,
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "git_commit": git_commit,
                "git_date": git_date,
                "processing_description": (
                    "Cible D6T brute + 55 features instantanées/dérivées/EWMA ; "
                    "temps STM32 retiré"
                ),
            }
        )
    records.sort(key=lambda item: item["nominal_start"])
    return records


def load_manual_values(path: Path) -> dict[str, dict[str, Any]]:
    """Read manual columns from a prior inventory before it is overwritten."""
    if not path.exists():
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook["Inventaire"]
        header_row = None
        header_map: dict[str, int] = {}
        for row in range(1, min(sheet.max_row, 20) + 1):
            values = [sheet.cell(row, column).value for column in range(1, sheet.max_column + 1)]
            if "Nom du fichier" in values:
                header_row = row
                header_map = {str(value): index + 1 for index, value in enumerate(values) if value}
                break
        if not header_row or "Nom du fichier" not in header_map:
            workbook.close()
            return {}
        result: dict[str, dict[str, Any]] = {}
        name_column = header_map["Nom du fichier"]
        for row in range(header_row + 1, sheet.max_row + 1):
            file_name = sheet.cell(row, name_column).value
            if not file_name:
                continue
            result[str(file_name)] = {
                column: sheet.cell(row, header_map[column]).value
                for column in MANUAL_COLUMNS
                if column in header_map
            }
        workbook.close()
        return result
    except Exception:
        return {}


def title_block(
    sheet,
    title: str,
    subtitle: str,
    last_column: int,
    generated_at: datetime,
    git_revision: str,
) -> None:
    last_letter = get_column_letter(last_column)
    sheet.merge_cells(f"A1:{last_letter}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells(f"A2:{last_letter}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Aptos", size=11, color=WHITE)
    sheet["A2"].fill = PatternFill("solid", fgColor=BLUE)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 30

    sheet.merge_cells(f"A3:{last_letter}3")
    sheet["A3"] = (
        f"Généré le {generated_at:%d/%m/%Y à %H:%M:%S} | "
        f"Révision Git : {git_revision} | Une ligne = un fichier"
    )
    sheet["A3"].font = Font(name="Aptos", size=9, italic=True, color="44546A")
    sheet["A3"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    sheet["A3"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 22


def add_category_band(sheet, row: int, groups: list[tuple[str, int, int, str]]) -> None:
    for label, start, stop, color in groups:
        sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=stop)
        cell = sheet.cell(row, start, label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in range(start, stop + 1):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=color)
            sheet.cell(row, column).border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[row].height = 22


def add_excel_table(
    sheet,
    name: str,
    header_row: int,
    end_row: int,
    end_column: int,
    start_column: int = 1,
) -> None:
    reference = (
        f"{get_column_letter(start_column)}{header_row}:"
        f"{get_column_letter(end_column)}{end_row}"
    )
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def style_header(sheet, row: int, columns: int) -> None:
    for cell in sheet[row][:columns]:
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY)
    sheet.row_dimensions[row].height = 56


def auto_width(sheet, min_width: int = 10, max_width: int = 42) -> None:
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        maximum = 0
        for row in range(1, min(sheet.max_row, 250) + 1):
            cell = sheet.cell(row, column)
            if cell.value is None:
                continue
            value = str(cell.value)
            maximum = max(maximum, max(len(line) for line in value.splitlines()))
        sheet.column_dimensions[letter].width = min(max(maximum + 2, min_width), max_width)


def apply_inventory_formats(sheet, header_row: int, manual_start_column: int) -> None:
    headers = {
        sheet.cell(header_row, column).value: column
        for column in range(1, sheet.max_column + 1)
    }
    for row in range(header_row + 1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 42
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Aptos", size=9, color=BLACK)
        for header, number_format in {
            "Date acquisition": "dd/mm/yyyy",
            "Heure début": "hh:mm:ss",
            "Fin estimée": "dd/mm/yyyy hh:mm:ss.000",
            "Durée": "[h]:mm:ss.000",
            "Durée moteur actif": "[h]:mm:ss.0",
            "Fréquence (Hz)": "0.000",
            "Période médiane (ms)": "0.000",
            "Consigne frein (A)": "0.00",
            "Part active (%)": "0.00",
        }.items():
            if header in headers:
                sheet.cell(row, headers[header]).number_format = number_format
        for header, column in headers.items():
            if isinstance(header, str) and (
                "(°C)" in header
                or "(tr/min)" in header
                or "Iq moteur" in header
                or "Écart max" in header
            ):
                sheet.cell(row, column).number_format = "0.000"
        for column in range(manual_start_column, sheet.max_column + 1):
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)

    if "Statut de validation" in headers and sheet.max_row > header_row:
        column_letter = get_column_letter(headers["Statut de validation"])
        validation = DataValidation(
            type="list",
            formula1='"À valider,Validé,Corrigé,Non applicable"',
            allow_blank=True,
        )
        validation.promptTitle = "Validation humaine"
        validation.prompt = "Choisir le statut après vérification des métadonnées."
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}{header_row + 1}:{column_letter}{sheet.max_row}")

    if "Confiance frein" in headers:
        letter = get_column_letter(headers["Confiance frein"])
        sheet.conditional_formatting.add(
            f"{letter}{header_row + 1}:{letter}{sheet.max_row}",
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("confirmer",{letter}{header_row + 1}))'],
                fill=PatternFill("solid", fgColor=YELLOW),
            ),
        )
    if "Qualité données" in headers:
        letter = get_column_letter(headers["Qualité données"])
        sheet.conditional_formatting.add(
            f"{letter}{header_row + 1}:{letter}{sheet.max_row}",
            FormulaRule(
                formula=[f'{letter}{header_row + 1}="OK"'],
                fill=PatternFill("solid", fgColor=LIGHT_GREEN),
            ),
        )


def write_simple_table(
    sheet,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    table_name: str,
    widths: dict[str, float] | None = None,
) -> int:
    for column, header in enumerate(headers, 1):
        sheet.cell(start_row, column, header)
    style_header(sheet, start_row, len(headers))
    for row_index, values in enumerate(rows, start_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=9, color=BLACK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row_index].height = 34
    end_row = start_row + max(1, len(rows))
    if rows:
        add_excel_table(sheet, table_name, start_row, end_row, len(headers))
    sheet.freeze_panes = f"A{start_row + 1}"
    auto_width(sheet)
    if widths:
        header_map = {header: index + 1 for index, header in enumerate(headers)}
        for header, width in widths.items():
            if header in header_map:
                sheet.column_dimensions[get_column_letter(header_map[header])].width = width
    return end_row


def populate_raw_inventory(
    workbook: Workbook,
    records: list[dict[str, Any]],
    manual: dict[str, dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.active
    sheet.title = "Inventaire"
    headers = [
        "ID session",
        "Nom du fichier",
        "Date acquisition",
        "Heure début",
        "Fin estimée",
        "Durée",
        "Lignes",
        "Fréquence (Hz)",
        "Qualité données",
        "Consigne vitesse (tr/min)",
        "Source vitesse",
        "Profil vitesse",
        "Vitesse active médiane (tr/min)",
        "Vitesse active P05 (tr/min)",
        "Vitesse active P95 (tr/min)",
        "Durée moteur actif",
        "Part active (%)",
        "Séquences moteur",
        "Mode frein",
        "Consigne frein (A)",
        "Libellé frein",
        "Source frein",
        "Confiance frein",
        "Iq moteur actif médian (A)",
        "Iq moteur actif P05 (A)",
        "Iq moteur actif P95 (A)",
        "D6T début (°C)",
        "D6T min (°C)",
        "D6T max (°C)",
        "D6T fin (°C)",
        "Étendue D6T (°C)",
        "DS18B20 début (°C)",
        "DS18B20 min (°C)",
        "DS18B20 max (°C)",
        "DS18B20 fin (°C)",
        "Étendue DS18B20 (°C)",
        "Profil thermique",
        "Fichier prétraité associé",
        "Correspondance",
        "Description automatique",
        "Remarques automatiques",
        *MANUAL_COLUMNS,
    ]
    title_block(
        sheet,
        "Inventaire des données brutes PMSM",
        "Périmètre : datalogging/logs — mesures, métadonnées inférées, contrôles et champs de validation équipe",
        len(headers),
        generated_at,
        git_revision,
    )
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(headers))
    sheet["A4"] = (
        "Important : la vitesse de consigne et le courant de frein ne sont pas enregistrés dans les CSV. "
        "La vitesse est inférée du plateau mesuré ; le frein provient du nom prétraité ou d'une inférence explicitement signalée."
    )
    sheet["A4"].fill = PatternFill("solid", fgColor=YELLOW)
    sheet["A4"].font = Font(name="Aptos", size=10, bold=True, color="7F6000")
    sheet["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 38
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(headers))
    sheet["A5"] = (
        "Les cellules jaunes en fin de ligne sont réservées aux compléments de l'équipe. "
        "Elles sont conservées lors d'une régénération du classeur."
    )
    sheet["A5"].font = Font(name="Aptos", size=9, italic=True, color="44546A")
    sheet["A5"].alignment = Alignment(wrap_text=True)

    add_category_band(
        sheet,
        7,
        [
            ("IDENTIFICATION & TEMPS", 1, 9, NAVY),
            ("VITESSE", 10, 18, BLUE),
            ("FREIN / CHARGE", 19, 26, "806000"),
            ("THERMIQUE", 27, 37, DARK_GREEN),
            ("LIENS & SYNTHÈSE", 38, 41, "5B5B5B"),
            ("À COMPLÉTER PAR L'ÉQUIPE", 42, len(headers), "9C6500"),
        ],
    )
    header_row = 8
    for column, header in enumerate(headers, 1):
        sheet.cell(header_row, column, header)
    style_header(sheet, header_row, len(headers))

    header_comments = {
        "Date acquisition": "Date issue du token YYYYMMDD_HHMMSS du nom, créé avec l'heure du PC.",
        "Durée": "(dernier stm32_time_ms - premier stm32_time_ms) / 1000.",
        "Consigne vitesse (tr/min)": "Non journalisée : valeur inférée du plateau de motor_speed_mech_rpm.",
        "Consigne frein (A)": "Courant du frein externe. Ne pas confondre avec motor_iq_a.",
        "D6T min (°C)": "Température IR cible : pixel sélectionné du capteur D6T.",
        "DS18B20 min (°C)": "Température externe de référence utilisée comme feature.",
    }
    for header, text in header_comments.items():
        column = headers.index(header) + 1
        sheet.cell(header_row, column).comment = Comment(text, "Codex")

    for index, record in enumerate(records, header_row + 1):
        prior = manual.get(record["file_name"], {})
        manual_values = [
            prior.get("Opérateur"),
            prior.get("Banc / moteur"),
            prior.get("Objet de l'essai (équipe)"),
            prior.get("Commentaires équipe"),
            prior.get("Statut de validation") or "À valider",
            prior.get("Mise à jour manuelle"),
        ]
        values = [
            record["session_id"],
            record["file_name"],
            record["nominal_start"].date(),
            record["nominal_start"].time(),
            record["nominal_end"],
            duration_value(record["duration_s"]),
            record["rows"],
            record["frequency_hz"],
            record["quality"],
            record["speed_setpoint"],
            record["speed_source"],
            record["speed_profile"],
            record["speed_median"],
            record["speed_p05"],
            record["speed_p95"],
            duration_value(record["active_duration_s"]),
            record["active_ratio"],
            record["sequence_count"],
            record["brake_mode"],
            record["brake_value"],
            record["brake_label"],
            record["brake_source"],
            record["brake_confidence"],
            record["iq_median"],
            record["iq_p05"],
            record["iq_p95"],
            record["d6_start"],
            record["d6_min"],
            record["d6_max"],
            record["d6_end"],
            record["d6_range"],
            record["ds_start"],
            record["ds_min"],
            record["ds_max"],
            record["ds_end"],
            record["ds_range"],
            record["thermal_profile"],
            record["processed_name"],
            record["pair_status"],
            record["description"],
            record["notes"],
            *manual_values,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(index, column, value)
        sheet.cell(index, 2).hyperlink = record["relative_path"]
        sheet.cell(index, 2).style = "Hyperlink"
        processed_path = f"pretraitement/logs_processed_ewma/{record['processed_name']}"
        sheet.cell(index, 38).hyperlink = processed_path
        sheet.cell(index, 38).style = "Hyperlink"

    add_excel_table(sheet, "InventaireBrut", header_row, header_row + len(records), len(headers))
    apply_inventory_formats(sheet, header_row, len(headers) - len(MANUAL_COLUMNS) + 1)
    sheet.freeze_panes = "C9"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = f"{header_row}:{header_row}"
    auto_width(sheet, min_width=11, max_width=38)
    for name in ["Nom du fichier", "Description automatique", "Remarques automatiques"]:
        sheet.column_dimensions[get_column_letter(headers.index(name) + 1)].width = 36
    for name in ["Source vitesse", "Source frein", "Profil thermique"]:
        sheet.column_dimensions[get_column_letter(headers.index(name) + 1)].width = 28
    for name in MANUAL_COLUMNS:
        sheet.column_dimensions[get_column_letter(headers.index(name) + 1)].width = 24


def populate_raw_summary(
    workbook: Workbook,
    records: list[dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Synthèse")
    title_block(
        sheet,
        "Synthèse des données brutes",
        "Vue rapide des 8 sessions actuellement présentes dans datalogging/logs",
        9,
        generated_at,
        git_revision,
    )
    metrics = [
        ("Fichiers", len(records), "CSV bruts inventoriés"),
        ("Période", f"{min(r['nominal_start'] for r in records):%d/%m/%Y} – {max(r['nominal_start'] for r in records):%d/%m/%Y}", "Selon les noms de fichiers"),
        ("Durée cumulée", duration_text(sum(r["duration_s"] for r in records)), "Somme des durées STM32"),
        ("Échantillons", sum(r["rows"] for r in records), "Toutes sessions"),
        ("Cadence", f"{np.median([r['frequency_hz'] for r in records]):.1f} Hz", "Médiane des fichiers"),
        ("D6T globale", f"{min(r['d6_min'] for r in records):.1f} à {max(r['d6_max'] for r in records):.1f} °C", "Cible IR"),
        ("Vitesse", "2 000 tr/min", "Inférée des plateaux mesurés"),
        ("Intégrité", f"{sum(r['quality'] == 'OK' for r in records)}/{len(records)} OK", "Aucune valeur manquante/infinie"),
    ]
    sheet["A5"] = "Indicateur"
    sheet["B5"] = "Valeur"
    sheet["C5"] = "Précision"
    style_header(sheet, 5, 3)
    for row, metric in enumerate(metrics, 6):
        for column, value in enumerate(metric, 1):
            sheet.cell(row, column, value)
            sheet.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[row].height = 28
    add_excel_table(sheet, "SyntheseBrutKPI", 5, 5 + len(metrics), 3)

    sheet["E5"] = "Profil frein"
    sheet["F5"] = "Fichiers"
    sheet["G5"] = "Durée cumulée"
    style_header(sheet, 5, 7)
    brake_groups: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        brake_groups.setdefault(record["brake_label"], []).append(record)
    for row, (label, group) in enumerate(sorted(brake_groups.items()), 6):
        sheet.cell(row, 5, label)
        sheet.cell(row, 6, len(group))
        sheet.cell(row, 7, duration_value(sum(item["duration_s"] for item in group)))
        sheet.cell(row, 7).number_format = "[h]:mm:ss.000"
    brake_end_row = 5 + len(brake_groups)
    add_excel_table(
        sheet, "SyntheseBrutFrein", 5, brake_end_row, 7, start_column=5
    )

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Nombre de fichiers par profil de frein"
    chart.y_axis.title = "Profil frein"
    chart.x_axis.title = "Fichiers"
    data = Reference(sheet, min_col=6, min_row=5, max_row=brake_end_row)
    categories = Reference(sheet, min_col=5, min_row=6, max_row=brake_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 6
    chart.width = 11
    sheet.add_chart(chart, "E12")

    sheet["A16"] = "Points à valider par l'équipe"
    sheet["A16"].font = Font(name="Aptos Display", size=14, bold=True, color=NAVY)
    caveats = [
        "Les trois profils « Mixte » sont des inférences solides mais restent à confirmer humainement.",
        "La consigne de vitesse 2 000 tr/min est inférée du plateau mesuré ; elle n'est pas stockée dans le log.",
        "Le courant motor_iq_a est une mesure moteur et ne constitue pas la consigne du frein externe.",
        "Compléter les colonnes jaunes (opérateur, banc, objet, commentaires et validation) avant publication définitive.",
    ]
    for row, item in enumerate(caveats, 17):
        sheet.cell(row, 1, "•")
        sheet.cell(row, 2, item)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 28
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 34
    sheet.column_dimensions["E"].width = 28
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 19
    sheet.sheet_view.showGridLines = False


def populate_sequences(
    workbook: Workbook,
    sequences: list[dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Séquences moteur")
    headers = [
        "ID session",
        "Nom du fichier",
        "Séquence",
        "Début relatif",
        "Fin relative",
        "Durée active",
        "Début nominal estimé",
        "Fin nominale estimée",
        "Vitesse moyenne (tr/min)",
        "Vitesse médiane (tr/min)",
        "Vitesse P05 (tr/min)",
        "Vitesse P95 (tr/min)",
        "Iq moyen (A)",
        "Iq médian (A)",
        "Iq P05 (A)",
        "Iq P95 (A)",
        "D6T début (°C)",
        "D6T fin (°C)",
        "D6T max (°C)",
        "DS18B20 début (°C)",
        "DS18B20 fin (°C)",
        "Profil frein session",
        "Confiance frein",
    ]
    title_block(
        sheet,
        "Séquences principales de marche moteur",
        "Détection : |vitesse| > 1 500 tr/min, phases actives ≥ 10 s ; interruptions ≤ 5 s fusionnées",
        len(headers),
        generated_at,
        git_revision,
    )
    rows = []
    for item in sequences:
        rows.append(
            [
                item["session_id"],
                item["file_name"],
                item["sequence"],
                duration_value(item["start_offset_s"]),
                duration_value(item["end_offset_s"]),
                duration_value(item["duration_s"]),
                item["nominal_start"],
                item["nominal_end"],
                item["speed_mean"],
                item["speed_median"],
                item["speed_p05"],
                item["speed_p95"],
                item["iq_mean"],
                item["iq_median"],
                item["iq_p05"],
                item["iq_p95"],
                item["d6_start"],
                item["d6_end"],
                item["d6_max"],
                item["ds_start"],
                item["ds_end"],
                item["brake_label"],
                item["brake_confidence"],
            ]
        )
    end_row = write_simple_table(sheet, 5, headers, rows, "SequencesMoteur")
    for row in range(6, end_row + 1):
        for column in [4, 5, 6]:
            sheet.cell(row, column).number_format = "[h]:mm:ss.0"
        for column in [7, 8]:
            sheet.cell(row, column).number_format = "dd/mm/yyyy hh:mm:ss.0"
        for column in range(9, 22):
            sheet.cell(row, column).number_format = "0.000"
    sheet.freeze_panes = "C6"
    sheet.sheet_view.showGridLines = False


def populate_raw_traceability(
    workbook: Workbook,
    records: list[dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Traçabilité")
    headers = [
        "ID session",
        "Nom du fichier",
        "Chemin relatif",
        "Taille (octets)",
        "Taille (MiB)",
        "SHA-256",
        "Dernier commit Git",
        "Date commit Git",
        "Premier stm32_time_ms",
        "Dernier stm32_time_ms",
        "Durée (s)",
        "Durée lisible",
        "Période min (ms)",
        "Période médiane (ms)",
        "Période max (ms)",
        "Trous > 2 périodes",
        "Intervalles non positifs",
        "Valeurs manquantes",
        "Valeurs infinies",
        "Doublons lignes",
        "Colonnes",
        "Schéma",
        "Détail qualité",
    ]
    title_block(
        sheet,
        "Traçabilité et intégrité des fichiers bruts",
        "Empreintes, tailles, provenance Git et contrôles temporels",
        len(headers),
        generated_at,
        git_revision,
    )
    rows = [
        [
            record["session_id"],
            record["file_name"],
            record["relative_path"],
            record["size_bytes"],
            record["size_bytes"] / 1024**2,
            record["sha256"],
            record["git_commit"],
            record["git_date"],
            record["first_stm32_ms"],
            record["last_stm32_ms"],
            record["duration_s"],
            duration_text(record["duration_s"]),
            record["period_min_ms"],
            record["period_ms"],
            record["period_max_ms"],
            record["time_gaps"],
            record["nonpositive_intervals"],
            record["missing"],
            record["infinite"],
            record["duplicate_rows"],
            record["columns"],
            "OK — 8 colonnes attendues" if record["columns"] == 8 else "À contrôler",
            record["quality_detail"],
        ]
        for record in records
    ]
    end_row = write_simple_table(sheet, 5, headers, rows, "TracabiliteBrut")
    for row in range(6, end_row + 1):
        sheet.cell(row, 5).number_format = "0.000"
        for column in [9, 10, 11, 13, 14, 15]:
            sheet.cell(row, column).number_format = "0.000"
    sheet.column_dimensions["F"].width = 68
    sheet.sheet_view.showGridLines = False


def populate_raw_dictionary(
    workbook: Workbook, generated_at: datetime, git_revision: str
) -> None:
    sheet = workbook.create_sheet("Dictionnaire données")
    headers = ["Position", "Colonne", "Unité", "Rôle", "Description", "Présente dans prétraité"]
    descriptions = [
        (1, "stm32_time_ms", "ms", "Temps", "Uptime de la carte STM32 ; sert à calculer durée et fréquence.", "Non"),
        (2, "d6t_temp_c", "°C", "Cible", "Température IR du pixel D6T sélectionné ; cible de la régression.", "Oui, colonne 1"),
        (3, "ds18b20_temp_c", "°C", "Feature", "Température externe de référence DS18B20.", "Oui"),
        (4, "motor_ud_v", "V", "Feature", "Tension moteur sur l'axe d.", "Oui"),
        (5, "motor_uq_v", "V", "Feature", "Tension moteur sur l'axe q.", "Oui"),
        (6, "motor_speed_mech_rpm", "tr/min", "Feature", "Vitesse mécanique mesurée ; ce n'est pas la consigne.", "Oui"),
        (7, "motor_id_a", "A", "Feature", "Courant moteur sur l'axe d.", "Oui"),
        (8, "motor_iq_a", "A", "Feature", "Courant moteur sur l'axe q ; ce n'est pas le courant de frein externe.", "Oui"),
    ]
    title_block(
        sheet,
        "Dictionnaire du format brut",
        "Ordre exact des huit colonnes CSV séparées par « ; »",
        len(headers),
        generated_at,
        git_revision,
    )
    write_simple_table(sheet, 5, headers, [list(row) for row in descriptions], "DictionnaireBrut")
    sheet.sheet_view.showGridLines = False


def populate_methodology(
    workbook: Workbook,
    kind: str,
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Méthodologie")
    headers = ["Thème", "Règle appliquée", "Provenance / limite"]
    common_rows = [
        ["ID session", "Token YYYYMMDD_HHMMSS extrait du nom.", "Clé de jointure brut ↔ prétraité."],
        ["Date/heure", "Horodatage nominal extrait du nom du fichier.", "Créé par datetime.now() côté dashboard ; ne pas utiliser LastWriteTime NTFS."],
        ["Durée", "(dernier - premier stm32_time_ms) / 1000.", "Mesurée depuis l'uptime STM32, pas depuis l'heure civile."],
        ["Fréquence", "1000 / médiane des différences positives de stm32_time_ms.", "10 Hz pour les huit sessions."],
        ["Consigne vitesse", "Plateau actif arrondi à la centaine la plus proche.", "Inférence très solide, mais la consigne CFG n'est pas enregistrée."],
        ["Marche moteur", "|motor_speed_mech_rpm| > 1 500 tr/min.", "Durée active estimée au pas médian."],
        ["Séquences", "Phases actives ≥ 10 s ; interruptions ≤ 5 s fusionnées.", "Évite de compter les transitoires isolés."],
        ["Frein fixe", "Valeur extraite du préfixe 0.05/0.10/0.15 du nom prétraité.", "Annotation de nom ; absence de télémétrie frein dans les CSV."],
        ["Frein mixte", "Nom prétraité sans préfixe + plusieurs paliers d'Iq moteur.", "Inférence à confirmer. motor_iq_a n'est pas le courant de frein."],
        ["Plage thermique", "Minimum et maximum calculés séparément pour D6T et DS18B20.", "D6T = cible IR ; DS18B20 = feature externe."],
        ["Périmètre", "Uniquement datalogging/logs et logs_processed_ewma.", "datalogging/raw_data (legacy, sans stm32_time_ms) est hors périmètre demandé."],
        ["Champs jaunes", "Réservés aux informations et validations humaines.", "Conservés si le générateur est relancé sur le classeur existant."],
    ]
    if kind == "processed":
        common_rows.extend(
            [
                ["Appariement", "Même ID session après retrait de l'éventuel préfixe frein.", "Chaque prétraité actuel possède exactement un brut source."],
                ["Validation", "Recalcul complet des 56 colonnes avec les règles du script courant.", "Tolérance 5,1×10⁻⁷, cohérente avec l'export à six décimales."],
                ["Temps", "stm32_time_ms volontairement absent du prétraité.", "Durée/date/fréquence du catalogue sont héritées du brut associé."],
                ["Nommage", "Le script de prétraitement réutilise normalement le nom du brut.", "Les cinq préfixes de frein ont été ajoutés après génération ; attention aux doublons lors d'une relance."],
            ]
        )
    title_block(
        sheet,
        "Méthodologie et limites",
        "Séparation explicite entre faits mesurés, métadonnées de nom et inférences",
        len(headers),
        generated_at,
        git_revision,
    )
    write_simple_table(sheet, 5, headers, common_rows, f"Methodologie{kind.title()}")
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 65
    sheet.column_dimensions["C"].width = 65
    sheet.sheet_view.showGridLines = False


def populate_processed_inventory(
    workbook: Workbook,
    records: list[dict[str, Any]],
    manual: dict[str, dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.active
    sheet.title = "Inventaire"
    headers = [
        "ID session",
        "Nom du fichier",
        "Fichier brut source",
        "Date acquisition",
        "Heure début",
        "Fin estimée",
        "Durée",
        "Lignes",
        "Colonnes totales",
        "Cible",
        "Features",
        "En-tête",
        "Temps STM32 inclus",
        "Fréquence source (Hz)",
        "Spans EWMA",
        "Consigne vitesse (tr/min)",
        "Source vitesse",
        "Vitesse active médiane (tr/min)",
        "Mode frein",
        "Consigne frein (A)",
        "Libellé frein",
        "Source frein",
        "Confiance frein",
        "D6T min (°C)",
        "D6T max (°C)",
        "Étendue D6T (°C)",
        "DS18B20 min (°C)",
        "DS18B20 max (°C)",
        "Étendue DS18B20 (°C)",
        "Lignes = brut",
        "Validation prétraitement",
        "Écart max absolu",
        "Cellules divergentes",
        "Valeurs manquantes / infinies",
        "Description prétraitement",
        "Description session automatique",
        "Remarques automatiques",
        *MANUAL_COLUMNS,
    ]
    title_block(
        sheet,
        "Inventaire des données prétraitées PMSM / EWMA",
        "Périmètre : pretraitement/logs_processed_ewma — une cible D6T + 55 features, avec rattachement au brut",
        len(headers),
        generated_at,
        git_revision,
    )
    sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(headers))
    sheet["A4"] = (
        "Les fichiers n'ont pas d'en-tête ni de timestamp : date, durée, températures et vitesse sont héritées du brut apparié. "
        "Les 56 colonnes ont été recomputées et contrôlées fichier par fichier."
    )
    sheet["A4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A4"].font = Font(name="Aptos", size=10, bold=True, color=NAVY)
    sheet["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 38
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(headers))
    sheet["A5"] = (
        "Attention : le générateur de prétraitement conserve normalement le nom brut. Les préfixes de frein ont été ajoutés après génération."
    )
    sheet["A5"].font = Font(name="Aptos", size=9, italic=True, color="9C5700")

    add_category_band(
        sheet,
        7,
        [
            ("IDENTIFICATION & TEMPS HÉRITÉS", 1, 8, NAVY),
            ("FORMAT / FEATURES", 9, 15, BLUE),
            ("VITESSE", 16, 18, "2F75B5"),
            ("FREIN", 19, 23, "806000"),
            ("THERMIQUE", 24, 29, DARK_GREEN),
            ("CONTRÔLES", 30, 35, "5B5B5B"),
            ("SYNTHÈSE & ÉQUIPE", 36, len(headers), "9C6500"),
        ],
    )
    header_row = 8
    for column, header in enumerate(headers, 1):
        sheet.cell(header_row, column, header)
    style_header(sheet, header_row, len(headers))
    sheet.cell(header_row, headers.index("Écart max absolu") + 1).comment = Comment(
        "Écart maximal entre le fichier et un recalcul complet avant arrondi à six décimales.",
        "Codex",
    )
    sheet.cell(header_row, headers.index("Consigne frein (A)") + 1).comment = Comment(
        "Valeur portée par le nom du fichier, jamais par les 56 colonnes numériques.",
        "Codex",
    )

    for index, record in enumerate(records, header_row + 1):
        prior = manual.get(record["file_name"], {})
        manual_values = [
            prior.get("Opérateur"),
            prior.get("Banc / moteur"),
            prior.get("Objet de l'essai (équipe)"),
            prior.get("Commentaires équipe"),
            prior.get("Statut de validation") or "À valider",
            prior.get("Mise à jour manuelle"),
        ]
        values = [
            record["session_id"],
            record["file_name"],
            record["raw_name"],
            record["nominal_start"].date(),
            record["nominal_start"].time(),
            record["nominal_end"],
            duration_value(record["duration_s"]),
            record["rows"],
            record["columns"],
            "d6t_temp_c (1)",
            f"{record['feature_columns']} prédicteurs",
            record["header"],
            record["time_included"],
            record["frequency_hz"],
            ", ".join(str(value) for value in record["spans"]),
            record["speed_setpoint"],
            record["speed_source"],
            record["speed_median"],
            record["brake_mode"],
            record["brake_value"],
            record["brake_label"],
            record["brake_source"],
            record["brake_confidence"],
            record["d6_min"],
            record["d6_max"],
            record["d6_range"],
            record["ds_min"],
            record["ds_max"],
            record["ds_range"],
            "Oui" if record["rows_match"] else "Non",
            record["validation"],
            record["max_abs_difference"],
            record["divergent_cells"],
            f"{record['missing']} / {record['infinite']}",
            record["processing_description"],
            record["description"],
            record["notes"],
            *manual_values,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(index, column, value)
        sheet.cell(index, 2).hyperlink = record["relative_path"]
        sheet.cell(index, 2).style = "Hyperlink"
        sheet.cell(index, 3).hyperlink = record["raw_path"].relative_to(ROOT).as_posix()
        sheet.cell(index, 3).style = "Hyperlink"

    add_excel_table(sheet, "InventairePretraite", header_row, header_row + len(records), len(headers))
    apply_inventory_formats(sheet, header_row, len(headers) - len(MANUAL_COLUMNS) + 1)
    sheet.freeze_panes = "D9"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = f"{header_row}:{header_row}"
    auto_width(sheet, min_width=11, max_width=40)
    for name in ["Nom du fichier", "Fichier brut source", "Validation prétraitement"]:
        sheet.column_dimensions[get_column_letter(headers.index(name) + 1)].width = 38
    for name in ["Description prétraitement", "Description session automatique", "Remarques automatiques"]:
        sheet.column_dimensions[get_column_letter(headers.index(name) + 1)].width = 40
    for name in MANUAL_COLUMNS:
        sheet.column_dimensions[get_column_letter(headers.index(name) + 1)].width = 24


def populate_processed_summary(
    workbook: Workbook,
    records: list[dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Synthèse")
    title_block(
        sheet,
        "Synthèse des données prétraitées EWMA",
        "Couverture, format et résultat des contrôles de correspondance",
        8,
        generated_at,
        git_revision,
    )
    metrics = [
        ("Fichiers", len(records), "Tous appariés à un brut"),
        ("Lignes cumulées", sum(r["rows"] for r in records), "Identiques aux sources"),
        ("Structure", "56 colonnes", "1 cible + 55 features"),
        ("En-tête / temps", "Non / absent", "Format NanoEdge actuel"),
        ("Fréquence source", "10 Hz", "Déduite des bruts"),
        ("Spans EWMA", "6600, 16800, 31800, 47400", "Adaptés de la référence 2 Hz"),
        ("Conformité", f"{sum(r['validation'].startswith('OK') for r in records)}/{len(records)} OK", "Recalcul complet à 10⁻⁶"),
        ("Valeurs invalides", sum(r["missing"] + r["infinite"] for r in records), "NaN + Inf"),
    ]
    rows = [[label, value, note] for label, value, note in metrics]
    write_simple_table(sheet, 5, ["Indicateur", "Valeur", "Précision"], rows, "SynthesePretraiteKPI")
    sheet["A16"] = "Convention de fichier"
    sheet["A16"].font = Font(name="Aptos Display", size=14, bold=True, color=NAVY)
    notes = [
        "Chaque fichier est un CSV numérique séparé par « ; », avec décimale « . » et six décimales.",
        "Colonne 1 : d6t_temp_c (cible non lissée). Colonnes 2 à 56 : 55 prédicteurs.",
        "Le détail positionnel complet figure dans l'onglet « Schéma 56 colonnes ».",
        "Les métadonnées temporelles du catalogue viennent du brut car stm32_time_ms a été retiré.",
        "Une régénération du prétraitement peut créer des noms sans préfixe frein : vérifier avant de synchroniser sur SharePoint.",
    ]
    for row, note in enumerate(notes, 17):
        sheet.cell(row, 1, "•")
        sheet.cell(row, 2, note)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[row].height = 28
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 34
    sheet.column_dimensions["C"].width = 38
    sheet.sheet_view.showGridLines = False


def processed_schema_rows() -> list[list[Any]]:
    units = {
        "d6t_temp_c": "°C",
        "ds18b20_temp_c": "°C",
        "motor_ud_v": "V",
        "motor_uq_v": "V",
        "motor_speed_mech_rpm": "tr/min",
        "motor_id_a": "A",
        "motor_iq_a": "A",
        "u_s": "V",
        "i_s": "A",
        "S_el": "VA",
        "speed_current": "tr/min·A",
        "speed_power": "tr/min·VA",
    }
    formulas = {
        "u_s": "sqrt(motor_ud_v² + motor_uq_v²)",
        "i_s": "sqrt(motor_id_a² + motor_iq_a²)",
        "S_el": "1,5 × u_s × i_s",
        "speed_current": "motor_speed_mech_rpm × i_s",
        "speed_power": "motor_speed_mech_rpm × S_el",
    }
    rows: list[list[Any]] = [
        [1, "d6t_temp_c", "d6t_temp_c", "Cible", units["d6t_temp_c"], "Instantanée", None, None, "Température IR non lissée"]
    ]
    position = 2
    for base in EWM_COLUMNS:
        role = "Dérivée" if base in DERIVED_COLUMNS else "Mesurée"
        rows.append(
            [position, base, base, "Prédicteur", units[base], f"{role}, instantanée", None, None, formulas.get(base, "Valeur brute")]
        )
        position += 1
        for span in [6600, 16800, 31800, 47400]:
            equivalent_s = span / 10.0
            rows.append(
                [
                    position,
                    f"{base}_ewma_{span}",
                    base,
                    "Prédicteur",
                    units[base],
                    "EWMA (adjust=False)",
                    span,
                    equivalent_s,
                    f"alpha = 2 / ({span} + 1)",
                ]
            )
            position += 1
    return rows


def populate_processed_schema(
    workbook: Workbook, generated_at: datetime, git_revision: str
) -> None:
    sheet = workbook.create_sheet("Schéma 56 colonnes")
    headers = [
        "Position CSV (1-based)",
        "Nom logique",
        "Variable de base",
        "Rôle",
        "Unité",
        "Transformation",
        "Span (échantillons)",
        "Span / fréquence (s)",
        "Formule / remarque",
    ]
    title_block(
        sheet,
        "Schéma positionnel des 56 colonnes",
        "Indispensable car les fichiers prétraités sont exportés sans ligne d'en-tête",
        len(headers),
        generated_at,
        git_revision,
    )
    rows = processed_schema_rows()
    end_row = write_simple_table(sheet, 5, headers, rows, "SchemaPretraite56")
    for row in range(6, end_row + 1):
        sheet.cell(row, 8).number_format = "0.0"
        if sheet.cell(row, 4).value == "Cible":
            for column in range(1, len(headers) + 1):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["I"].width = 42
    sheet.sheet_view.showGridLines = False


def populate_processed_controls(
    workbook: Workbook,
    records: list[dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Contrôles")
    headers = [
        "ID session",
        "Fichier prétraité",
        "Brut source",
        "Lignes prétraitées",
        "Lignes brutes",
        "Même nombre de lignes",
        "Forme attendue",
        "Valeurs manquantes",
        "Valeurs infinies",
        "Écart max absolu",
        "Cellules > 5,1E-7",
        "Résultat",
    ]
    title_block(
        sheet,
        "Contrôles brut ↔ prétraité",
        "Recalcul fichier par fichier des valeurs instantanées, dérivées et EWMA",
        len(headers),
        generated_at,
        git_revision,
    )
    rows = [
        [
            record["session_id"],
            record["file_name"],
            record["raw_name"],
            record["rows"],
            record["raw_rows"],
            "Oui" if record["rows_match"] else "Non",
            "Oui" if record["shape_match"] else "Non",
            record["missing"],
            record["infinite"],
            record["max_abs_difference"],
            record["divergent_cells"],
            record["validation"],
        ]
        for record in records
    ]
    end_row = write_simple_table(sheet, 5, headers, rows, "ControlesPretraitement")
    for row in range(6, end_row + 1):
        sheet.cell(row, 10).number_format = "0.0000000000E+00"
        if str(sheet.cell(row, 12).value).startswith("OK"):
            for column in range(1, len(headers) + 1):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    sheet.column_dimensions["L"].width = 48
    sheet.sheet_view.showGridLines = False


def populate_processed_traceability(
    workbook: Workbook,
    records: list[dict[str, Any]],
    generated_at: datetime,
    git_revision: str,
) -> None:
    sheet = workbook.create_sheet("Traçabilité")
    headers = [
        "ID session",
        "Nom du fichier",
        "Chemin relatif",
        "Taille (octets)",
        "Taille (MiB)",
        "SHA-256 prétraité",
        "SHA-256 brut source",
        "Dernier commit Git",
        "Date commit Git",
        "Séparateur",
        "Décimale",
        "Précision export",
        "En-tête",
        "Timestamp inclus",
        "Colonnes",
        "Cible + features",
    ]
    title_block(
        sheet,
        "Traçabilité des fichiers prétraités",
        "Empreintes d'intégrité, provenance et paramètres de sérialisation",
        len(headers),
        generated_at,
        git_revision,
    )
    rows = [
        [
            record["session_id"],
            record["file_name"],
            record["relative_path"],
            record["size_bytes"],
            record["size_bytes"] / 1024**2,
            record["sha256"],
            record["raw_sha256"],
            record["git_commit"],
            record["git_date"],
            record["separator"],
            record["decimal"],
            record["precision"],
            record["header"],
            record["time_included"],
            record["columns"],
            f"{record['target_columns']} + {record['feature_columns']}",
        ]
        for record in records
    ]
    end_row = write_simple_table(sheet, 5, headers, rows, "TracabilitePretraite")
    for row in range(6, end_row + 1):
        sheet.cell(row, 5).number_format = "0.000"
    sheet.column_dimensions["F"].width = 68
    sheet.column_dimensions["G"].width = 68
    sheet.sheet_view.showGridLines = False


def set_workbook_properties(workbook: Workbook, title: str, generated_at: datetime) -> None:
    workbook.properties.title = title
    workbook.properties.subject = "Catalogue historique des acquisitions PMSM"
    workbook.properties.creator = "STMicroelectronics / Codex"
    workbook.properties.description = (
        "Inventaire auto-documenté des fichiers de datalogging PMSM, avec provenance des métadonnées."
    )
    workbook.properties.created = generated_at.replace(tzinfo=None)
    workbook.properties.modified = generated_at.replace(tzinfo=None)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def build_workbooks() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    generated_at = datetime.now().astimezone()
    git_revision = git_value("rev-parse", "--short", "HEAD")
    raw_manual = load_manual_values(RAW_WORKBOOK)
    processed_manual = load_manual_values(PROCESSED_WORKBOOK)

    raw_records, sequences = analyse_raw_files()
    processed_records = analyse_processed_files(raw_records)

    raw_workbook = Workbook()
    set_workbook_properties(raw_workbook, "Inventaire des données brutes PMSM", generated_at)
    populate_raw_inventory(raw_workbook, raw_records, raw_manual, generated_at, git_revision)
    populate_raw_summary(raw_workbook, raw_records, generated_at, git_revision)
    populate_sequences(raw_workbook, sequences, generated_at, git_revision)
    populate_raw_traceability(raw_workbook, raw_records, generated_at, git_revision)
    populate_raw_dictionary(raw_workbook, generated_at, git_revision)
    populate_methodology(raw_workbook, "raw", generated_at, git_revision)
    raw_workbook.save(RAW_WORKBOOK)

    processed_workbook = Workbook()
    set_workbook_properties(
        processed_workbook, "Inventaire des données prétraitées PMSM / EWMA", generated_at
    )
    populate_processed_inventory(
        processed_workbook,
        processed_records,
        processed_manual,
        generated_at,
        git_revision,
    )
    populate_processed_summary(
        processed_workbook, processed_records, generated_at, git_revision
    )
    populate_processed_schema(processed_workbook, generated_at, git_revision)
    populate_processed_controls(
        processed_workbook, processed_records, generated_at, git_revision
    )
    populate_processed_traceability(
        processed_workbook, processed_records, generated_at, git_revision
    )
    populate_methodology(processed_workbook, "processed", generated_at, git_revision)
    processed_workbook.save(PROCESSED_WORKBOOK)
    return raw_records, processed_records


def main() -> None:
    raw_records, processed_records = build_workbooks()
    print(f"Créé : {RAW_WORKBOOK.name} ({len(raw_records)} fichiers)")
    print(f"Créé : {PROCESSED_WORKBOOK.name} ({len(processed_records)} fichiers)")


if __name__ == "__main__":
    main()
