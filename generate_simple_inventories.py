"""Create lightweight Excel summaries from the two detailed inventories."""

from __future__ import annotations

from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
RAW_SOURCE = ROOT / "inventaire_donnees_brutes.xlsx"
PROCESSED_SOURCE = ROOT / "inventaire_donnees_pretraitees_ewma.xlsx"
RAW_OUTPUT = ROOT / "resume_simple_donnees_brutes.xlsx"
PROCESSED_OUTPUT = ROOT / "resume_simple_donnees_pretraitees_ewma.xlsx"

NAVY = "003B5C"
BLUE = "0077B6"
WHITE = "FFFFFF"
LIGHT_BLUE = "DDEBF7"
YELLOW = "FFF2CC"


def as_datetime(date_value, time_value) -> datetime:
    if isinstance(date_value, datetime):
        date_value = date_value.date()
    if isinstance(time_value, datetime):
        time_value = time_value.time()
    if not isinstance(time_value, time):
        time_value = time(0, 0)
    return datetime.combine(date_value, time_value)


def temperature_ranges(d6_minimum, d6_maximum, ds_minimum, ds_maximum) -> str:
    d6 = f"{float(d6_minimum):.1f}–{float(d6_maximum):.1f}".replace(".", ",")
    ds = f"{float(ds_minimum):.2f}–{float(ds_maximum):.2f}".replace(".", ",")
    return f"D6T {d6} °C | DS18B20 {ds} °C"


def read_inventory(path: Path) -> tuple[list[str], list[dict]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Inventaire"]
    headers = [sheet.cell(8, column).value for column in range(1, sheet.max_column + 1)]
    rows = []
    for row_index in range(9, sheet.max_row + 1):
        row = {
            header: sheet.cell(row_index, column).value
            for column, header in enumerate(headers, 1)
        }
        if row.get("Nom du fichier"):
            rows.append(row)
    workbook.close()
    return headers, rows


def create_summary(
    output: Path,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list],
    table_name: str,
    widths: list[float],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Résumé"
    last_column = get_column_letter(len(headers))

    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Aptos", size=10, color=WHITE)
    sheet["A2"].fill = PatternFill("solid", fgColor=BLUE)
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 30

    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = "Vue volontairement simplifiée — consulter l'inventaire détaillé pour les méthodes, contrôles et valeurs complètes."
    sheet["A3"].font = Font(name="Aptos", size=9, italic=True, color=NAVY)
    sheet["A3"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sheet["A3"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[3].height = 26

    header_row = 5
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 36

    for row_index, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row_index, 2 if len(headers) == 7 else 3).number_format = "dd/mm/yyyy hh:mm:ss"
        sheet.cell(row_index, 3 if len(headers) == 7 else 4).number_format = "[h]:mm:ss.000"
        brake_column = 6 if len(headers) == 7 else 7
        if str(sheet.cell(row_index, brake_column).value).startswith("Mixte"):
            sheet.cell(row_index, brake_column).fill = PatternFill("solid", fgColor=YELLOW)
        sheet.row_dimensions[row_index].height = 34

    end_row = header_row + len(rows)
    table = Table(displayName=table_name, ref=f"A{header_row}:{last_column}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "B6"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = "5:5"

    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    workbook.properties.title = title
    workbook.properties.subject = "Résumé simple des fichiers PMSM"
    workbook.properties.creator = "STMicroelectronics / Codex"
    workbook.save(output)


def main() -> None:
    _, raw_rows = read_inventory(RAW_SOURCE)
    raw_values = []
    for row in raw_rows:
        brake = row["Libellé frein"]
        if str(brake).startswith("Mixte"):
            brake = "Mixte (à confirmer)"
        raw_values.append(
            [
                row["Nom du fichier"],
                as_datetime(row["Date acquisition"], row["Heure début"]),
                row["Durée"],
                temperature_ranges(
                    row["D6T min (°C)"],
                    row["D6T max (°C)"],
                    row["DS18B20 min (°C)"],
                    row["DS18B20 max (°C)"],
                ),
                f"{int(row['Consigne vitesse (tr/min)']):,} tr/min".replace(",", " "),
                brake,
                row["Profil thermique"],
            ]
        )

    create_summary(
        RAW_OUTPUT,
        "Résumé simple — données brutes PMSM",
        "Une ligne par fichier : date, durée et conditions principales de l'essai",
        [
            "Nom du fichier",
            "Date et heure",
            "Durée",
            "Températures min–max",
            "Vitesse estimée",
            "Frein",
            "Type d'enregistrement",
        ],
        raw_values,
        "ResumeSimpleBrut",
        [34, 21, 17, 40, 20, 22, 42],
    )

    _, processed_rows = read_inventory(PROCESSED_SOURCE)
    processed_values = []
    for row in processed_rows:
        brake = row["Libellé frein"]
        if str(brake).startswith("Mixte"):
            brake = "Mixte (à confirmer)"
        processed_values.append(
            [
                row["Nom du fichier"],
                row["Fichier brut source"],
                as_datetime(row["Date acquisition"], row["Heure début"]),
                row["Durée"],
                temperature_ranges(
                    row["D6T min (°C)"],
                    row["D6T max (°C)"],
                    row["DS18B20 min (°C)"],
                    row["DS18B20 max (°C)"],
                ),
                f"{int(row['Consigne vitesse (tr/min)']):,} tr/min".replace(",", " "),
                brake,
                "Données pour IA : 1 température cible + 55 variables prétraitées EWMA",
            ]
        )

    create_summary(
        PROCESSED_OUTPUT,
        "Résumé simple — données prétraitées PMSM / EWMA",
        "Une ligne par fichier : source brute, conditions principales et contenu IA",
        [
            "Nom du fichier",
            "Fichier brut associé",
            "Date et heure",
            "Durée",
            "Températures min–max",
            "Vitesse estimée",
            "Frein",
            "Contenu",
        ],
        processed_values,
        "ResumeSimplePretraite",
        [38, 34, 21, 17, 40, 20, 22, 52],
    )

    print(f"Créé : {RAW_OUTPUT.name} ({len(raw_values)} fichiers)")
    print(f"Créé : {PROCESSED_OUTPUT.name} ({len(processed_values)} fichiers)")


if __name__ == "__main__":
    main()
